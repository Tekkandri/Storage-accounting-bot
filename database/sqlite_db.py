import sqlite3 as sq
from datetime import datetime, timedelta
import openpyxl
from config import tz
import pandas as pd

#----------actions with text file----------
def get_balance() -> int:
    with open("balance.txt","r") as file:
        return int(file.readline())

def set_new_balance(balance: int):
    with open("balance.txt","w") as file:
        file.write(str(balance))
#----------------------------------------

def start_sql():
    global db, cur
    db = sq.connect('./database/sklad.db')
    cur = db.cursor()
    if db:
        print("Database connected successfully")
    db.commit()

def set_balance_hist_line(sum):
    cur.execute("INSERT INTO balance VALUES(?,?)", (sum, datetime.now(tz=tz)))
    db.commit()

def get_balance_hist(month: int, year: int) -> dict:
    cur.execute("SELECT * FROM balance")
    history = cur.fetchall()
    result = {
        "data": ["\n", 0]
    }
    book = openpyxl.Workbook()
    sheet = book.active
    i = 0
    for line in history:
        if datetime.strptime(line[1].split('.')[0],"%Y-%m-%d %H:%M:%S").month == month and \
                datetime.strptime(line[1].split('.')[0],"%Y-%m-%d %H:%M:%S").year == year:
            result["data"][0] += f"✏ {line[0]} руб. -- {line[1].split('.')[0]}\n"
            result["data"][1] += line[0]
            i += 1
            j = 1
            for col in line:
                cell = sheet.cell(row=i, column=j)
                cell.value = col
                j += 1
    book.save(f"Баланс {datetime.now(tz=tz).date()}.xlsx")
    return result

def set_new_purchase(category, count, sum, auto,bool):
    cur.execute("INSERT INTO purchases VALUES(?,?,?,?,?,?)",(category, count, sum, auto, datetime.now(tz=tz),bool))
    db.commit()

def get_day_item_stat() -> dict:
    cur.execute("SELECT * from purchases")
    items = cur.fetchall()
    correct_items = []
    book = openpyxl.Workbook()
    sheet = book.active
    result = {}
    cur.execute("SELECT category FROM prices")
    categories = cur.fetchall()
    for cat in categories:
        result[cat[0]] = [0,0]
    for item in items:
        item_date = datetime.strptime(item[4].split('.')[0],"%Y-%m-%d %H:%M:%S")
        if datetime.now(tz=tz) - item_date.now(tz=tz) < timedelta(days=3) and datetime.now(tz=tz).year == item_date.year:
            correct_items.append(item)
            if item[5] and item[0] in result.keys():
                result[item[0]][1] += item[2]
                result[item[0]][0] += item[1]

    i = 0
    for item in correct_items:
        i += 1
        j = 1
        for col in item[:-1]:
            cell = sheet.cell(row=i, column=j)
            cell.value = col
            j += 1
    book.save(f"Дневная статистика {datetime.now(tz=tz).date()}.xlsx")

    return result

def get_month_item_stat(month: int, year: int) -> dict:
    cur.execute("SELECT * from purchases")
    items = cur.fetchall()
    correct_items = []
    book = openpyxl.Workbook()
    sheet = book.active
    result = {}
    cur.execute("SELECT category FROM prices")
    categories = cur.fetchall()
    for cat in categories:
        result[cat[0]] = [0, 0]
    for item in items:
        item_date = datetime.strptime(item[4].split('.')[0],"%Y-%m-%d %H:%M:%S")
        if month == item_date.month and year == item_date.year:
            correct_items.append(item)
            if item[5] and item[0] in result.keys():
                result[item[0]][1] += item[2]
                result[item[0]][0] += item[1]

    i = 0
    for item in correct_items:
        i += 1
        j = 1
        for col in item[:-1]:
            cell = sheet.cell(row=i, column=j)
            cell.value = col
            j += 1
    book.save(f"Месячная статистика {datetime.now(tz=tz).date()}.xlsx")

    return result

def set_new_price(cat: str, price: int):
    cur.execute(f"UPDATE prices SET price='{price}' WHERE category='{cat}'")
    db.commit()

def get_price(cat: str) -> int:
    cur.execute(f"SELECT price FROM prices WHERE category='{cat}'")
    return cur.fetchone()[0]

def get_all():
    cur.execute("SELECT * FROM purchases")
    items = cur.fetchall()
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.cell(row=1, column=1).value = "category"
    sheet.cell(row=1, column=2).value = "count"
    sheet.cell(row=1, column=3).value = "sum"
    sheet.cell(row=1, column=4).value = "auto"
    sheet.cell(row=1, column=5).value = "date"
    sheet.cell(row=1, column=6).value = "bool"
    i = 1
    for item in items:
        i += 1
        j = 1
        for col in item:
            cell = sheet.cell(row=i, column=j)
            cell.value = col
            j += 1
    book.save(f"Выгрузка.xlsx")


def get_all_balance_hist():
    cur.execute("SELECT * FROM balance")
    items = cur.fetchall()
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.cell(row=1, column=1).value = "sum"
    sheet.cell(row=1, column=2).value = "date"
    i = 1
    for item in items:
        i += 1
        j = 1
        for col in item:
            cell = sheet.cell(row=i, column=j)
            cell.value = col
            j += 1
    book.save(f"Выгрузка баланс.xlsx")

def set_new_balance_db(filename: str):
    wb = pd.read_excel(filename)
    wb.to_sql(name='balance', con=db, if_exists='replace')
    cur.execute("PRAGMA foreign_keys = 0;")
    cur.execute("CREATE TABLE sqlitestudio_temp_table AS SELECT * FROM balance;")
    cur.execute("DROP TABLE balance;")
    cur.execute("""CREATE TABLE balance (
        sum      INT,
        date     DATE
    );""")
    cur.execute("""INSERT INTO balance (
                              sum,
                              date
                          )
                          SELECT sum,
                                 date
                            FROM sqlitestudio_temp_table;""")
    cur.execute("DROP TABLE sqlitestudio_temp_table")
    cur.execute("PRAGMA foreign_keys = 1;")
    db.commit()

def set_new_db(filename: str):
    cur.execute("SELECT * FROM purchases")
    old_db = cur.fetchall()
    wb = pd.read_excel(filename)
    lst = []
    for index, row in wb.iterrows():
        lst.append((row["category"],row["count"],row["sum"],row["auto"],row["date"],row["bool"]))
    for line in old_db:
        if line not in lst:
            set_balance_hist_line(line[2])
            balance = get_balance()
            set_new_balance(balance+line[2])

    wb.to_sql(name='purchases', con=db, if_exists='replace')
    cur.execute("PRAGMA foreign_keys = 0;")
    cur.execute("CREATE TABLE sqlitestudio_temp_table AS SELECT * FROM purchases;")
    cur.execute("DROP TABLE purchases;")
    cur.execute("""CREATE TABLE purchases (
    category TEXT,
    count    INTEGER,
    sum      INTEGER,
    auto     TEXT,
    date     TEXT,
    bool     BOOLEAN
);""")
    cur.execute("""INSERT INTO purchases (
                          category,
                          count,
                          sum,
                          auto,
                          date,
                          bool
                      )
                      SELECT category,
                             count,
                             sum,
                             auto,
                             date,
                             bool
                        FROM sqlitestudio_temp_table;""")
    cur.execute("DROP TABLE sqlitestudio_temp_table")
    cur.execute("PRAGMA foreign_keys = 1;")
    db.commit()

def add_category(category: str):
    cur.execute("INSERT INTO prices VALUES(?,?)",(category, 0))
    db.commit()

def get_categories() -> list:
    cur.execute("SELECT category FROM prices")
    data = cur.fetchall()
    return data

def delete_category(category: str):
    cur.execute(f"DELETE FROM prices WHERE category='{category}'")
    db.commit()